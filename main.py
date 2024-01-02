import os
from dotenv import load_dotenv
from flask import Flask, jsonify, request
import plaid
from plaid.api import plaid_api
import pandas as pd
from datetime import datetime
from plaid.model.transactions_get_request import TransactionsGetRequest
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi

# Load environment variables
load_dotenv()

PLAID_CLIENT_ID = os.getenv('PLAID_CLIENT_ID')
PLAID_SECRET = os.getenv('PLAID_DEVELOPMENT_SECRET')
# ACCESS_TOKEN is now dynamically obtained after Plaid Link process
MONGODB_USER = os.getenv('MONGODB_USER')
MONGODB_PW = os.getenv('MONGODB_PW')

# MongoDB URI
uri = f"mongodb+srv://{MONGODB_USER}:{MONGODB_PW}@cluster0.mlyhyf1.mongodb.net/?retryWrites=true&w=majority"
# Create MongoDB client
mongodb_client = MongoClient(uri, server_api=ServerApi('1'))

db = mongodb_client['plaid-integration']  # Replace with your MongoDB database name
access_tokens_collection = db.access_tokens

# Constants
HEADERS = ["TRANS_ID", "DATE", "DESCRIPTION", "INSTITUTION", "ACCOUNT TYPE", "CATEGORY", "AMOUNT", "STATUS"]

app = Flask(__name__)

# Initialize Plaid client
def init_plaid_client():
    config = plaid.Configuration(
        host=plaid.Environment.Development,
        api_key={
            'clientId': PLAID_CLIENT_ID,
            'secret': PLAID_SECRET
        }
    )
    client = plaid.ApiClient(config)
    return plaid_api.PlaidApi(client)

plaid_client = init_plaid_client()

@app.route('/create_link_token', methods=['POST'])
def create_link_token():
    try:
        response = plaid_client.link_token_create({
            'user': {
                'client_user_id': 'unique_user_id',  # Adjust as needed
            },
            'client_name': "Plaid Integration App",
            'products': ["transactions"],
            'country_codes': ["US"],
            'language': "en",
        })
        return jsonify(response.to_dict())
    except plaid.ApiException as e:
        return jsonify({'error': str(e)})

@app.route('/get_access_token', methods=['POST'])
def get_access_token():
    public_token = request.json['public_token']
    try:
        exchange_response = plaid_client.item_public_token_exchange({'public_token': public_token})
        access_token = exchange_response['access_token']

        # Insert access token into MongoDB
        access_tokens_collection.insert_one({'access_token': access_token})

        return jsonify({'access_token': access_token})
    except plaid.ApiException as e:
        return jsonify({'error': str(e)})

def get_stored_access_token():
    token_document = access_tokens_collection.find_one()  # Adjust query as needed
    return token_document['access_token'] if token_document else None


if __name__ == '__main__':
    app.run(port=int(os.environ.get("PORT", 5000)), host='0.0.0.0', debug=True)

# Fetch transactions from Plaid
def get_transactions(client, access_token, start_date_str, end_date_str):
    try:
        # Convert string dates to datetime objects in format 'YYYY-MM-DD'
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        # Create a TransactionsGetRequest object with the date objects
        request = TransactionsGetRequest(access_token=access_token, start_date=start_date, end_date=end_date)
        response = client.transactions_get(request)
        return response['transactions']
    except plaid.ApiException as e:
        print(f"Error fetching transactions: {e}")
        return []

# Format transactions into DataFrame
def format_transactions(transactions):
    formatted = []
    for txn in transactions:
        formatted.append({
            "TRANS_ID": txn.get('transaction_id', ''),
            "DATE": txn.get('date', ''),
            "DESCRIPTION": txn.get('name', ''),
            "INSTITUTION": 'TBD',  # Placeholder for actual institution data
            "ACCOUNT TYPE": 'TBD',  # Placeholder for account type
            "CATEGORY": ' > '.join(txn.get('category', [])),
            "AMOUNT": txn.get('amount', 0),
            "STATUS": 'Pending' if txn.get('pending', False) else 'Completed'
        })
    return pd.DataFrame(formatted, columns=HEADERS)

# Update Excel workbook with transaction data
def update_workbook(df, path):
    if df.empty:
        print("No transactions to update.")
        return

    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Transactions', index=False)
        ws = writer.book['Transactions']
        apply_styles_to_sheet(ws)

# Apply styles to Excel sheet
def apply_styles_to_sheet(ws):
    set_column_widths(ws)
    currency_style = NamedStyle(name='currency', number_format='"$"#,##0.00')
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = '@'  # TRANS_ID as text
        row[1].number_format = 'MM/DD/YYYY'  # DATE
        row[6].style = currency_style  # AMOUNT

def set_column_widths(ws):
    column_widths = {'A': 12, 'B': 11, 'C': 75, 'D': 13, 'E': 13, 'F': 38, 'G': 9, 'H': 9}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

# Reorder sheets in workbook
def reorder_sheets(path, sheet_name, index):
    book = load_workbook(path)
    sheet = book[sheet_name]
    book.remove(sheet)
    book._sheets.insert(index, sheet)
    book.save(path)
